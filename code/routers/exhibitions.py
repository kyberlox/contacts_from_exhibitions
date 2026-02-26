# routers/exhibitions.py
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, asc, or_, func
from typing import List, Optional
from datetime import date, datetime
import os
from pathlib import Path

from models.database import get_db
from models.exhibition import Exhibition
from models.user import User
from models.contact import Contact
from models.file import File as FileModel
from schemas import (
    ExhibitionCreate,
    ExhibitionUpdate,
    Exhibition as ExhibitionSchema,
    ExhibitionWithContacts,
    ExhibitionFilter,
    PaginationParams,
    ExhibitionSimple,
    ExhibitionWithContactsSimple
)
from schemas.base import PaginatedResponse
from services.auth import require_admin, require_auth, get_current_user

router = APIRouter(prefix="/exhibitions", tags=["Выставки"])

# Конфигурация
UPLOAD_DIR = Path("uploads/exhibitions")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

@router.post("/", response_model=ExhibitionSchema, status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_admin)])
async def create_exhibition(
        exhibition_data: ExhibitionCreate,
        db: AsyncSession = Depends(get_db)
):
    """Создание новой выставки"""
    # Убрана проверка на перекрывающиеся даты

    # Создаем выставку
    db_exhibition = Exhibition(**exhibition_data.dict(exclude_none=True))
    db.add(db_exhibition)
    await db.commit()
    await db.refresh(db_exhibition)

    # Возвращаем данные вручную
    return {
        "id": db_exhibition.id,
        "title": db_exhibition.title,
        "description": db_exhibition.description,
        "start_date": db_exhibition.start_date,
        "end_date": db_exhibition.end_date,
        "preview_file_id": db_exhibition.preview_file_id,
        "created_at": db_exhibition.created_at,
        "updated_at": db_exhibition.updated_at,
    }

@router.get("/", response_model=PaginatedResponse, dependencies=[Depends(require_auth)])
async def get_exhibitions(
        pagination: PaginationParams = Depends(),
        active_only: bool = Query(False, description="Только активные выставки"),
        sort_by: str = Query("start_date", description="Поле для сортировки"),
        sort_desc: bool = Query(True, description="Сортировка по убыванию"),
        db: AsyncSession = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Получение списка выставок с пагинацией и сортировкой"""

    # Строим базовый запрос
    query = select(Exhibition)

    # Фильтр по активным выставкам
    if active_only:
        today = date.today()
        query = query.where(
            (Exhibition.start_date <= today) &
            (Exhibition.end_date >= today)
        )

    if not current_user.is_admin:
        # Получаем ID выставок, где пользователь создавал контакты
        subquery = (
            select(Contact.exhibition_id)
            .where(Contact.author_id == current_user.id)
            .distinct()
        )

        # Фильтруем: активные ИЛИ те, где пользователь участвовал
        query = query.where(
            or_(
                Exhibition.is_active == True,
                Exhibition.id.in_(subquery)
            )
        )

    # Сортировка
    if hasattr(Exhibition, sort_by):
        order_field = getattr(Exhibition, sort_by)
        query = query.order_by(desc(order_field) if sort_desc else asc(order_field))
    else:
        # Сортировка по дате начала по умолчанию
        query = query.order_by(desc(Exhibition.start_date))

    # Общее количество
    total_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(total_query)
    total = total_result.scalar()

    # Пагинация
    query = query.offset(pagination.skip).limit(pagination.limit)

    # Выполняем запрос
    result = await db.execute(query)
    exhibitions = result.scalars().all()

    # Преобразуем вручную, чтобы избежать проблем с from_orm и связанными объектами
    items = []
    for exhibition in exhibitions:
        items.append({
            "id": exhibition.id,
            "title": exhibition.title,
            "description": exhibition.description,
            "start_date": exhibition.start_date,
            "end_date": exhibition.end_date,
            "preview_file_id": exhibition.preview_file_id,
            "created_at": exhibition.created_at,
            "updated_at": exhibition.updated_at,
        })

    return PaginatedResponse(
        total=total,
        skip=pagination.skip,
        limit=pagination.limit,
        items=items
    )



@router.get("/{exhibition_id}", response_model=ExhibitionWithContactsSimple, dependencies=[Depends(require_admin)])
async def get_exhibition_id(
        exhibition_id: int,
        db: AsyncSession = Depends(get_db)
):
    """Получение выставки по ID с контактами"""
    # Получаем выставку
    result = await db.execute(
        select(Exhibition).where(Exhibition.id == exhibition_id)
    )
    exhibition = result.scalar_one_or_none()

    if not exhibition:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Выставка не найдена"
        )

    # Получаем контакты для этой выставки
    from models.contact import Contact
    contacts_result = await db.execute(
        select(Contact).where(Contact.exhibition_id == exhibition_id)
    )
    contacts = contacts_result.scalars().all()

    # Формируем список контактов в простом формате
    contacts_list = []
    for contact in contacts:
        contacts_list.append({
            "id": contact.id,
            "title": contact.title,
            "full_name": contact.full_name,
            "position": contact.position,
            "email": contact.email,
            "phone_number": contact.phone_number,
            "created_at": contact.created_at.isoformat() if contact.created_at else None,
        })

    # Формируем ответ
    response_data = {
        "id": exhibition.id,
        "title": exhibition.title,
        "description": exhibition.description,
        "start_date": exhibition.start_date,
        "end_date": exhibition.end_date,
        "preview_file_id": exhibition.preview_file_id,
        "created_at": exhibition.created_at,
        "updated_at": exhibition.updated_at,
        "contacts": contacts_list
    }

    # Для отладки
    print(f"Returning exhibition data: {response_data}")

    return response_data

@router.put("/{exhibition_id}", response_model=ExhibitionSchema, dependencies=[Depends(require_admin)])
async def update_exhibition(
        exhibition_id: int,
        exhibition_data: ExhibitionUpdate,
        db: AsyncSession = Depends(get_db)
):
    """Обновление выставки"""
    result = await db.execute(
        select(Exhibition).where(Exhibition.id == exhibition_id)
    )
    exhibition = result.scalar_one_or_none()

    if not exhibition:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Выставка не найдена"
        )

    # Обновляем поля
    update_data = exhibition_data.dict(exclude_unset=True)
    for field, value in update_data.items():
        if value is not None:
            setattr(exhibition, field, value)

    await db.commit()
    await db.refresh(exhibition)

    return exhibition

@router.delete("/{exhibition_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(require_admin)])
async def delete_exhibition(
        exhibition_id: int,
        db: AsyncSession = Depends(get_db)
):
    """Удаление выставки"""
    result = await db.execute(
        select(Exhibition).where(Exhibition.id == exhibition_id)
    )
    exhibition = result.scalar_one_or_none()

    if not exhibition:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Выставка не найдена"
        )

    # TODO: Проверить, есть ли связанные контакты
    # Можно добавить каскадное удаление или предупреждение

    await db.delete(exhibition)
    await db.commit()

    return None

@router.post("/{exhibition_id}/preview", response_model=ExhibitionSchema, dependencies=[Depends(require_admin)])
async def upload_exhibition_preview(
        exhibition_id: int,
        file: UploadFile = File(...),
        db: AsyncSession = Depends(get_db)
):
    """Загрузка превью для выставки"""
    # Проверяем существование выставки
    result = await db.execute(
        select(Exhibition).where(Exhibition.id == exhibition_id)
    )
    exhibition = result.scalar_one_or_none()

    if not exhibition:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Выставка не найдена"
        )

    # Проверяем размер файла
    file.file.seek(0, 2)  # Переходим в конец файла
    file_size = file.file.tell()
    file.file.seek(0)  # Возвращаемся в начало

    if file_size > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Файл слишком большой. Максимальный размер: {MAX_FILE_SIZE // 1024 // 1024}MB"
        )

    # Проверяем расширение файла
    file_extension = Path(file.filename).suffix.lower()
    if file_extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Недопустимый формат файла. Разрешены: {', '.join(ALLOWED_IMAGE_EXTENSIONS)}"
        )

    # Генерируем уникальное имя файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_filename = f"{exhibition_id}_{timestamp}{file_extension}"
    file_path = UPLOAD_DIR / unique_filename

    # Сохраняем файл
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)

    # Создаем запись о файле в БД
    db_file = FileModel(
        name=file.filename,
        format=file_extension.lstrip('.'),
        path=str(file_path),
        url=f"/uploads/exhibitions/{unique_filename}"
    )
    db.add(db_file)
    await db.flush()  # Получаем ID файла

    # Обновляем выставку
    exhibition.preview_file_id = db_file.id
    await db.commit()
    await db.refresh(exhibition)

    # Возвращаем данные вручную в формате схемы
    return {
        "id": exhibition.id,
        "title": exhibition.title,
        "description": exhibition.description,
        "start_date": exhibition.start_date,
        "end_date": exhibition.end_date,
        "preview_file_id": exhibition.preview_file_id,
        "created_at": exhibition.created_at,
        "updated_at": exhibition.updated_at,
    }

@router.get("/{exhibition_id}/stats", dependencies=[Depends(require_admin)])
async def get_exhibition_stats(
        exhibition_id: int,
        db: AsyncSession = Depends(get_db)
):
    """Получение статистики по выставке"""
    from sqlalchemy import func
    from openpyxl import Workbook, load_workbook
    import io
    import requests
    import json
    try:
        # Проверяем существование выставки
        result = await db.execute(
            select(Exhibition).where(Exhibition.id == exhibition_id)
        )
        exhibition = result.scalar_one_or_none()

        if not exhibition:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Выставка не найдена"
            )

        # Получаем количество контактов
        from models.contact import Contact
        stmt_contacts = await db.execute(
            select(Contact).where(Contact.exhibition_id == exhibition_id)     # , Contact.is_validated == True
        )
        exhibition_contacts = stmt_contacts.scalars().all()

        if not exhibition_contacts:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Контакты по выставке {exhibition.title!r} не найдены"
            )
        
        wb = load_workbook('./statistics_pattern.xlsx')

        # Получаем статистику по типам контактов
        # wb = Workbook() 
        ws = wb.active
        ws.title = f"Статистика по выставке {exhibition.title}"
        # Запись данных
        # ws['A1'] = 'Город'
        # ws['B1'] = 'ФИО'
        # ws['C1'] = 'Должность'
        # ws['D1'] = 'Телефон'
        # ws['E1'] = 'Почта'
        # ws['F1'] = 'Дата и место пересечения'

        for i, contact in enumerate(exhibition_contacts, start=2):
            ws[f'A{i}'] = contact.city if contact.city else 'Не указан'
            ws[f'B{i}'] = contact.full_name if contact.full_name else 'Не указан'
            ws[f'C{i}'] = contact.position if contact.position else 'Не указана'
            ws[f'D{i}'] = contact.phone_number if contact.phone_number else 'Не указан'
            ws[f'E{i}'] = contact.email if contact.email else 'Не указан'
            ws[f'F{i}'] = f"{exhibition.start_date} {exhibition.title}"

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return StreamingResponse(excel_buffer,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=statistics_exhibition.xlsx"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )
